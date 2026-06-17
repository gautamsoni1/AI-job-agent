"""
Sheet Export Service — generates the two downloadable spreadsheets the
pipeline needs: (1) every job found before applying, (2) application
results after apply-all / manual apply.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import get_settings

settings = get_settings()


class SheetExportService:
    def __init__(self):
        self.output_dir = Path(settings.GENERATED_FILES_PATH) / "sheets"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_jobs_before_apply(self, pipeline_id: str, jobs: list[dict]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs Found"
        headers = [
            "Company", "Role", "Location", "Match Score", "Relevance Score",
            "Required Skills", "Salary Range", "Experience Required",
            "Employment Type", "Work Type", "Source", "Apply Link",
        ]
        self._header(ws, headers)
        for row_idx, job in enumerate(jobs, start=2):
            scout = job.get("scout_report", {}) or {}
            skills_gap = scout.get("required_vs_user_skills", {}) or {}
            required_skills = job.get("required_skills") or (
                skills_gap.get("matching_skills", []) + skills_gap.get("missing_skills", [])
            )
            values = [
                job.get("company", "") or "Company not provided",
                job.get("title", "") or "Role not provided",
                job.get("location", "") or "Location not provided",
                job.get("match_score", ""),
                scout.get("relevance_score", ""),
                ", ".join(required_skills),
                job.get("salary_range", "") or "Salary not provided",
                job.get("experience_required", "") or "Not specified",
                job.get("employment_type", "") or "Not specified",
                job.get("work_type", "") or "Not specified",
                job.get("source", "") or "Not specified",
                job.get("apply_link", "") or "Not provided",
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        self._autofit(ws, headers)
        path = self.output_dir / f"jobs_before_apply_{pipeline_id}.xlsx"
        wb.save(str(path))
        return str(path)

    def export_application_results(self, pipeline_id: str, results: list[dict]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Application Results"
        headers = ["Company", "Role", "Job ID", "Application Status", "Reason", "Applied At", "Apply Link"]
        self._header(ws, headers)
        for row_idx, r in enumerate(results, start=2):
            values = [
                r.get("company", "") or "Company not provided",
                r.get("title", "") or "Role not provided",
                r.get("job_id", ""),
                r.get("status", ""),
                r.get("reason", ""),
                r.get("applied_at", ""),
                r.get("apply_link", ""),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:
                    status_upper = str(value).upper()
                    if status_upper == "APPLIED":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif status_upper == "MANUAL_APPLY_REQUIRED":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif status_upper == "FAILED":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self._autofit(ws, headers)
        path = self.output_dir / f"application_results_{pipeline_id}.xlsx"
        wb.save(str(path))
        return str(path)

    def _header(self, ws, headers: list[str]):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    def _autofit(self, ws, headers: list[str]):
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22